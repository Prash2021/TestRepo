import openpyxl

wk= openpyxl.load_workbook("C:\\Users\\Prashanth\\PycharmProjects\\RobotAutomation\\Testdata\\testdata.xlsx")
sh= wk["Sheet1"]

rows=sh.max_row
columns= sh.max_column
print("Total no of rows are " +str(rows))
print("Total no of colums are" +str(columns))

# for i in range (1, rows+1):
#     for j in range (1, columns+1):
#         c=sh.cell(i,j)
#         print(c.value)

for i in sh['A1':'C5']:
    for c in i:
        print(c.value)
